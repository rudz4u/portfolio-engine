#!/usr/bin/env python3
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / 'Docs' / 'Investment Management.xlsx'
OUT_JSON = ROOT / 'Docs' / 'parsed_xlsx.json'
OUT_SUMMARY = ROOT / 'Docs' / 'parsed_xlsx_summary.md'

def sheet_to_list(sheet):
    rows = list(sheet.rows)
    if not rows:
        return []
    headers = [cell.value if cell.value is not None else f'col_{i}' for i, cell in enumerate(rows[0])]
    data = []
    for r in rows[1:]:
        row = {}
        for h, c in zip(headers, r):
            # preserve formulas as strings if present
            val = c.value
            row[h] = val
        data.append(row)
    return data

def main():
    wb = load_workbook(filename=str(XLSX_PATH), data_only=False)
    out = {}
    summary_lines = []
    summary_lines.append(f'# Parsed XLSX: {XLSX_PATH.name}\n')
    summary_lines.append('Sheets found:')
    for name in wb.sheetnames:
        sheet = wb[name]
        data = sheet_to_list(sheet)
        out[name] = data
        summary_lines.append(f'- {name}: {len(data)} data rows, {len(sheet[1]) if sheet.max_row>0 else 0} columns')
        # check for formulas in sheet
        formula_count = 0
        sample_formulas = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            for cell in row:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    formula_count += 1
                    if len(sample_formulas) < 5:
                        sample_formulas.append(cell.value)
        summary_lines.append(f'  - formulas found: {formula_count}')
        if sample_formulas:
            summary_lines.append('  - sample formulas:')
            for f in sample_formulas:
                summary_lines.append(f'    - `{f}`')

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    with open(OUT_SUMMARY, 'w', encoding='utf-8') as f:
        f.write('\n'.join(summary_lines))

    print(f'Wrote {OUT_JSON} and {OUT_SUMMARY}')

if __name__ == '__main__':
    main()
